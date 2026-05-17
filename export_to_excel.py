# -*- coding: utf-8 -*-
"""
导出为Excel格式（xlsx），完美支持换行符和长文本
"""
import csv

# 尝试导入openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️ 未安装 openpyxl，正在尝试安装...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

print("=" * 70)
print("📊 导出为Excel格式")
print("=" * 70)

# 读取数据
print("\n📂 读取CSV...")
with open('data/裁判文书_完整版_500条.csv', 'r', encoding='utf-8-sig') as f:
    rows = list(csv.DictReader(f))
print(f"✅ 读取 {len(rows)} 条记录")

# 创建Excel工作簿
print("\n📝 创建Excel工作簿...")
wb = Workbook()
ws = wb.active
ws.title = "裁判文书"

# 设置表头
headers = ['序号', '标题', '法院', '案号', '日期', '性别', '判决结果', '链接', '正文']
ws.append(headers)

# 设置表头样式
for cell in ws[1]:
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 写入数据
print("✍️ 写入数据...")
for i, row in enumerate(rows, 1):
    ws.append([
        row.get('序号', ''),
        row.get('标题', ''),
        row.get('法院', ''),
        row.get('案号', ''),
        row.get('日期', ''),
        row.get('性别', ''),
        row.get('判决结果', ''),
        row.get('链接', ''),
        row.get('正文', '')
    ])
    
    if i % 100 == 0:
        print(f"  已写入 {i}/{len(rows)}")

# 设置列宽
print("\n🎨 设置格式...")
ws.column_dimensions['A'].width = 8   # 序号
ws.column_dimensions['B'].width = 40  # 标题
ws.column_dimensions['C'].width = 20  # 法院
ws.column_dimensions['D'].width = 25  # 案号
ws.column_dimensions['E'].width = 12  # 日期
ws.column_dimensions['F'].width = 8   # 性别
ws.column_dimensions['G'].width = 50  # 判决结果
ws.column_dimensions['H'].width = 15  # 链接
ws.column_dimensions['I'].width = 80  # 正文

# 设置自动换行
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# 保存
output_file = 'data/裁判文书_500条.xlsx'
print(f"\n💾 保存到: {output_file}")
wb.save(output_file)
print("✅ 保存成功！")

print("\n" + "=" * 70)
print("🎉 完成！")
print("=" * 70)
print(f"\n生成文件: {output_file}")
print("说明：Excel格式完美支持换行符，可以直接用Excel打开")
print()


